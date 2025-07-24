from flask import Flask, render_template, jsonify, request
import json
import time
import threading
from datetime import datetime, timedelta
import os
import random
import pandas as pd
import glob
from openpyxl import load_workbook

app = Flask(__name__)

# 订单数据存储
orders_db = []
order_counter = 1
# 记录前端操作的时间戳，防止被Excel数据覆盖
frontend_operations = {}  # {order_id: last_operation_time}
# Excel文件监控变量
excel_file_modified_time = None  # 记录Excel文件最后修改时间
is_excel_updating = False  # 标记Excel是否正在被外部程序更新

# 订单状态常量
PENDING = 2          # 未完成
COMPLETED = 5        # 已完成

# Excel文件路径配置
EXCEL_FOLDER = os.path.join(os.path.expanduser("~"), "Desktop")  # 从桌面读取Excel文件
EXCEL_PATTERN = "*.xlsx"  # Excel文件匹配模式

def ensure_orders_folder():
    """确保桌面路径存在"""
    if not os.path.exists(EXCEL_FOLDER):
        print(f"❌ 桌面路径不存在: {EXCEL_FOLDER}")
        return False
    print(f"✅ 桌面路径正常: {EXCEL_FOLDER}")
    return True

def ensure_excel_files_writable():
    """确保Excel文件可写"""
    try:
        # 查找所有Excel文件
        excel_files = glob.glob(os.path.join(EXCEL_FOLDER, EXCEL_PATTERN))
        
        if not excel_files:
            print("📁 未找到Excel文件，跳过权限检查")
            return
        
        print("🔧 检查Excel文件写入权限...")
        
        for file_path in excel_files:
            try:
                # 检查文件是否可写
                if not os.access(file_path, os.W_OK):
                    print(f"⚠️  文件不可写: {os.path.basename(file_path)}")
                    
                    # 尝试修复文件权限
                    import stat
                    current_attrs = os.stat(file_path).st_mode
                    
                    # 添加写入权限
                    new_attrs = current_attrs | stat.S_IWRITE
                    os.chmod(file_path, new_attrs)
                    
                    print(f"✅ 已修复文件权限: {os.path.basename(file_path)}")
                else:
                    print(f"✅ 文件可写: {os.path.basename(file_path)}")
                    
            except Exception as e:
                print(f"❌ 处理文件权限时出错 {os.path.basename(file_path)}: {e}")
        
        print("🎉 Excel文件权限检查完成")
        
    except Exception as e:
        print(f"❌ 检查Excel文件权限时出错: {e}")

def map_order_status(status_text):
    """映射订单状态文本到数字状态"""
    # 保持原始状态，不进行映射转换
    # 这样前端可以显示原始的订单状态文本
    return status_text

def read_excel_orders():
    """从Excel文件读取订单数据"""
    global orders_db, excel_file_modified_time, is_excel_updating
    
    try:
        # 确保桌面路径存在
        if not ensure_orders_folder():
            print("❌ 无法访问桌面路径，使用空订单列表")
            orders_db = []
            return
        
        # 确保Excel文件可写
        ensure_excel_files_writable()
        
        # 查找所有Excel文件
        excel_files = glob.glob(os.path.join(EXCEL_FOLDER, EXCEL_PATTERN))
        
        if not excel_files:
            print("未找到Excel文件，使用空订单列表")
            orders_db = []
            return
        
        # 读取最新的Excel文件
        latest_file = max(excel_files, key=os.path.getctime)
        
        # 检查Excel文件是否被外部程序修改
        current_modified_time = os.path.getmtime(latest_file)
        
        if excel_file_modified_time is not None and current_modified_time > excel_file_modified_time:
            # Excel文件被外部程序修改了
            print(f"🔄 检测到Excel文件被外部程序修改: {latest_file}")
            is_excel_updating = True
            
            # 等待一段时间，确保外部程序完成写入
            time.sleep(3)
            
            # 再次检查修改时间，确保写入完成
            final_modified_time = os.path.getmtime(latest_file)
            if final_modified_time == current_modified_time:
                print("✅ Excel文件写入完成，开始读取新数据")
                # 保持冲突检测状态5秒，防止用户操作冲突
                time.sleep(5)
                is_excel_updating = False
            else:
                print("⏳ Excel文件仍在被修改，等待完成...")
                time.sleep(3)
                is_excel_updating = False
        
        # 更新文件修改时间
        excel_file_modified_time = current_modified_time
        
        print(f"读取Excel文件: {latest_file}")
        
        # 读取Excel数据
        df = pd.read_excel(latest_file, engine='openpyxl')
        print(f"成功读取Excel文件，数据行数: {len(df)}")
        
        # 保存现有订单数据用于比较
        old_orders_db = orders_db.copy()
        # 清空现有订单数据
        orders_db = []
        
        # 处理每一行数据
        valid_order_id = 1
        for index, row in df.iterrows():
            try:
                # 检查是否为空行
                order_number = row.get('订单编号')
                if pd.isna(order_number):
                    continue
                
                # 检查是否有前端操作记录
                has_frontend_operation = valid_order_id in frontend_operations
                frontend_operation = frontend_operations.get(valid_order_id)
                
                # 获取订单状态
                status_text = str(row.get('订单状态', '备货中'))
                
                # 跳过已取消的订单
                if status_text == '已取消':
                    continue
                
                # 根据您指定的列名映射数据
                order = {
                    'id': valid_order_id,
                    'number': str(order_number),
                    'status': map_order_status(status_text),
                    'userName': str(row.get('姓名', '未知')),
                    'phone': str(row.get('手机号码', '未知')),
                    'address': str(row.get('部门', '未知')),
                    'amount': float(row.get('订单金额', 0)),
                    'remark': f"取餐码: {str(row.get('取餐码', ''))}",
                    'orderTime': str(row.get('订单时间', datetime.now().isoformat())),
                    'dishes': []
                }
                
                # 处理商品信息（Unnamed: 39列）
                if 'Unnamed: 39' in row and pd.notna(row['Unnamed: 39']):
                    dishes_str = str(row['Unnamed: 39'])
                    if dishes_str and dishes_str != 'nan':
                        # 如果包含逗号，按逗号分割；否则作为单个商品
                        if ',' in dishes_str:
                            order['dishes'] = [{'name': dish.strip(), 'price': 0} for dish in dishes_str.split(',')]
                        else:
                            order['dishes'] = [{'name': dishes_str.strip(), 'price': 0}]
                        print(f"📦 订单{valid_order_id}商品信息: {dishes_str}")
                
                # 如果有前端操作记录，检查是否需要保护前端操作
                if has_frontend_operation and frontend_operation:
                    # 查找旧订单数据中的状态
                    existing_order = next((o for o in old_orders_db if o['id'] == valid_order_id), None)
                    if existing_order:
                        excel_status = order['status']
                        current_status = existing_order['status']
                        expected_status = frontend_operation['new_status']
                        
                        # 如果Excel状态与期望的前端状态不同，保持前端状态
                        if excel_status != expected_status:
                            print(f"🛡️  订单{valid_order_id}前端操作保护: Excel={excel_status}, 期望={expected_status}, 保持前端状态")
                            order['status'] = expected_status
                        else:
                            # 状态一致，但不要立即清除前端操作记录，等待一段时间
                            operation_time = frontend_operation['timestamp']
                            time_diff = datetime.now() - operation_time
                            
                            # 如果前端操作时间超过5分钟，才清除记录
                            if time_diff.total_seconds() > 300:  # 5分钟
                                del frontend_operations[valid_order_id]
                                print(f"✅ 订单{valid_order_id}状态同步且操作时间超过5分钟，清除前端操作记录")
                            else:
                                print(f"⏳ 订单{valid_order_id}状态同步，但操作时间较短({time_diff.total_seconds():.0f}秒)，保持保护")
                    else:
                        # 新订单，清除前端操作记录
                        del frontend_operations[valid_order_id]
                        print(f"✅ 新订单{valid_order_id}，清除前端操作记录")
                else:
                    # 没有前端操作记录，检查是否需要保护现有状态
                    existing_order = next((o for o in old_orders_db if o['id'] == valid_order_id), None)
                    if existing_order:
                        excel_status = order['status']
                        current_status = existing_order['status']
                        
                        # 如果现有订单状态与Excel状态不同，保持现有状态
                        if excel_status != current_status:
                            print(f"🛡️  订单{valid_order_id}状态保护: Excel={excel_status}, 内存={current_status}, 保持内存状态")
                            order['status'] = current_status
                            
                            # 如果状态被保护，重新记录前端操作
                            if valid_order_id not in frontend_operations:
                                frontend_operations[valid_order_id] = {
                                    'timestamp': datetime.now(),
                                    'old_status': excel_status,
                                    'new_status': current_status,
                                    'protected': True
                                }
                                print(f"📝 重新记录保护操作: 订单{valid_order_id} {excel_status}→{current_status}")
                
                orders_db.append(order)
                valid_order_id += 1
                
            except Exception as e:
                print(f"处理第{index + 1}行数据时出错: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        print(f"成功读取 {len(orders_db)} 个订单")
        
    except Exception as e:
        print(f"读取Excel文件时出错: {e}")
        import traceback
        traceback.print_exc()
        # 如果读取失败，保持现有数据不变

def background_excel_reader():
    """后台Excel读取线程"""
    while True:
        try:
            read_excel_orders()
            print(f"Excel数据刷新完成，当前订单数量: {len(orders_db)}")
        except Exception as e:
            print(f"后台Excel读取出错: {e}")
        
        # 等待1分钟
        time.sleep(60)

def get_orders_by_status(status=None):
    """根据状态获取订单"""
    if status is None:
        return orders_db
    
    # 处理字符串状态和数字状态的兼容性
    if isinstance(status, str):
        # 字符串状态直接匹配
        return [order for order in orders_db if order['status'] == status]
    else:
        # 数字状态匹配
        return [order for order in orders_db if order['status'] == status]

def update_order_status(order_id, new_status):
    """更新订单状态"""
    global orders_db, frontend_operations, is_excel_updating
    
    # 检查Excel是否正在被外部程序更新
    if is_excel_updating:
        print(f"⚠️  订单{order_id}状态更新被拒绝：Excel文件正在被外部程序更新")
        return False, "系统繁忙，请稍后再试"
    
    # 更新订单状态
    for order in orders_db:
        if order['id'] == order_id:
            old_status = order['status']
            order['status'] = new_status
            # 记录前端操作时间戳和旧状态
            frontend_operations[order_id] = {
                'timestamp': datetime.now(),
                'old_status': old_status,
                'new_status': new_status
            }
            print(f"📝 记录前端操作: 订单{order_id} {old_status}→{new_status}")
            # 同步更新Excel文件
            update_excel_order_status(order_id, new_status)
            return True, "操作成功"
    return False, "订单不存在"

def update_excel_order_status(order_id, new_status):
    """更新Excel文件中的订单状态"""
    try:
        # 查找最新的Excel文件
        excel_files = glob.glob(os.path.join(EXCEL_FOLDER, EXCEL_PATTERN))
        if not excel_files:
            print("未找到Excel文件，无法更新状态")
            return False
        
        latest_file = max(excel_files, key=os.path.getctime)
        print(f"更新Excel文件: {latest_file}")
        
        # 检查文件是否可写
        if not os.access(latest_file, os.W_OK):
            print(f"⚠️  Excel文件无写入权限，请检查文件是否被占用或设置为只读")
            print(f"   文件路径: {latest_file}")
            print(f"   建议操作:")
            print(f"   1. 关闭可能打开该文件的Excel程序")
            print(f"   2. 右键文件 -> 属性 -> 取消勾选'只读'")
            print(f"   3. 以管理员身份运行程序")
            return False
        
        # 使用openpyxl加载工作簿
        workbook = load_workbook(latest_file)
        worksheet = workbook.active
        
        # 找到对应的订单行（需要跳过空行）
        df = pd.read_excel(latest_file, engine='openpyxl')
        valid_rows = []
        
        for index, row in df.iterrows():
            order_number = row.get('订单编号')
            if not pd.isna(order_number):
                valid_rows.append(index + 2)  # +2 因为Excel从1开始且有标题行
        
        if order_id <= len(valid_rows):
            row_number = valid_rows[order_id - 1]  # 订单ID从1开始，转换为0基索引
            print(f"🔍 订单ID {order_id} 映射到Excel行号 {row_number}")
        else:
            print(f"订单ID {order_id} 超出有效订单范围")
            return False
        
        if row_number <= worksheet.max_row:
            # 处理状态更新
            if isinstance(new_status, str):
                # 如果是字符串状态，直接使用
                status_text = new_status
            else:
                # 如果是数字状态，进行映射
                status_mapping = {
                    PENDING: '备货中',
                    COMPLETED: '已完成'
                }
                status_text = status_mapping.get(new_status, '备货中')
            
            print(f"🔧 更新Excel: 订单{order_id}, 状态{new_status} -> 文本状态'{status_text}'")
            
            # 更新状态列（订单状态是第18列，R列）
            status_cell = worksheet.cell(row=row_number, column=18)
            status_cell.value = status_text
            
            # 保存文件
            workbook.save(latest_file)
            print(f"✅ 成功更新Excel文件，订单{order_id}状态改为{status_text}")
            return True
        else:
            print(f"订单ID {order_id} 超出Excel文件范围")
            return False
            
    except PermissionError as e:
        print(f"❌ Excel文件权限错误: {e}")
        print(f"   请确保Excel文件未被其他程序打开，且具有写入权限")
        return False
    except Exception as e:
        print(f"❌ 更新Excel文件时出错: {e}")
        return False

# Flask路由定义
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/orders')
def api_orders():
    """获取未完成订单"""
    # 获取所有非"已完成"状态的订单
    pending_orders = [order for order in orders_db if order['status'] != '已完成']
    return jsonify({
        'code': 1,
        'msg': 'success',
        'data': pending_orders
    })

@app.route('/api/orders/all')
def api_all_orders():
    """获取所有订单"""
    return jsonify({
        'code': 1,
        'msg': 'success',
        'data': orders_db
    })

@app.route('/api/order/<int:order_id>/<action>', methods=['POST'])
def api_update_order(order_id, action):
    """更新订单状态"""
    try:
        if action == 'complete':
            success, message = update_order_status(order_id, '已完成')
            if success:
                return jsonify({
                    'code': 1, 
                    'msg': f'{message}，Excel文件已同步更新'
                })
            else:
                return jsonify({
                    'code': 0, 
                    'msg': message
                })
        else:
            return jsonify({'code': 0, 'msg': '无效的操作'})
            
    except Exception as e:
        return jsonify({'code': 0, 'msg': f'操作失败: {str(e)}'})

@app.route('/api/statistics')
def api_statistics():
    """获取统计信息"""
    total_orders = len(orders_db)
    # 统计未完成订单（非"已完成"状态的订单）
    pending_orders = len([order for order in orders_db if order['status'] != '已完成'])
    # 统计已完成订单
    completed_orders = len([order for order in orders_db if order['status'] == '已完成'])
    total_amount = sum(order['amount'] for order in orders_db)
    
    return jsonify({
        'code': 1,
        'msg': 'success',
        'data': {
            'total_orders': total_orders,
            'pending_orders': pending_orders,
            'completed_orders': completed_orders,
            'total_amount': round(total_amount, 2)
        }
    })

@app.route('/api/orders/<status>')
def api_orders_by_status(status):
    """根据状态获取订单"""
    try:
        status_code = int(status)
        filtered_orders = get_orders_by_status(status_code)
        return jsonify({
            'code': 1,
            'msg': 'success',
            'data': filtered_orders
        })
    except ValueError:
        return jsonify({'code': 0, 'msg': '无效的状态码'})

@app.route('/api/excel-info')
def api_excel_info():
    """获取Excel文件信息"""
    try:
        if not ensure_orders_folder():
            return jsonify({
                'code': 0,
                'msg': '无法访问桌面路径',
                'data': None
            })
        excel_files = glob.glob(os.path.join(EXCEL_FOLDER, EXCEL_PATTERN))
        
        if excel_files:
            latest_file = max(excel_files, key=os.path.getctime)
            file_info = {
                'file_name': os.path.basename(latest_file),
                'file_path': latest_file,
                'last_modified': datetime.fromtimestamp(os.path.getmtime(latest_file)).isoformat(),
                'file_size': os.path.getsize(latest_file),
                'order_count': len(orders_db)
            }
        else:
            file_info = {
                'file_name': '无文件',
                'file_path': '',
                'last_modified': '',
                'file_size': 0,
                'order_count': 0
            }
        
        return jsonify({
            'code': 1,
            'msg': 'success',
            'data': file_info
        })
        
    except Exception as e:
        return jsonify({'code': 0, 'msg': f'获取文件信息失败: {str(e)}'})

@app.route('/api/excel-status')
def api_excel_status():
    """获取Excel文件状态信息"""
    try:
        if not ensure_orders_folder():
            return jsonify({
                'code': 0,
                'msg': '无法访问桌面路径',
                'data': None
            })
        excel_files = glob.glob(os.path.join(EXCEL_FOLDER, EXCEL_PATTERN))
        
        if excel_files:
            latest_file = max(excel_files, key=os.path.getctime)
            # 读取Excel文件获取最新状态
            df = pd.read_excel(latest_file, engine='openpyxl')
            
            # 统计各状态数量
            status_counts = df['订单状态'].value_counts().to_dict()
            
            return jsonify({
                'code': 1,
                'msg': 'success',
                'data': {
                    'file_name': os.path.basename(latest_file),
                    'last_modified': datetime.fromtimestamp(os.path.getmtime(latest_file)).isoformat(),
                    'status_counts': status_counts,
                    'total_orders': len(df)
                }
            })
        else:
            return jsonify({
                'code': 0,
                'msg': '未找到Excel文件',
                'data': None
            })
        
    except Exception as e:
        return jsonify({'code': 0, 'msg': f'获取Excel状态失败: {str(e)}'})

@app.route('/api/frontend-operations')
def api_frontend_operations():
    """获取前端操作记录"""
    try:
        operations_info = {}
        for order_id, operation_data in frontend_operations.items():
            if isinstance(operation_data, dict):
                operations_info[order_id] = {
                    'operation_time': operation_data['timestamp'].isoformat(),
                    'time_ago': str(datetime.now() - operation_data['timestamp']),
                    'old_status': operation_data['old_status'],
                    'new_status': operation_data['new_status']
                }
            else:
                # 兼容旧格式
                operations_info[order_id] = {
                    'operation_time': operation_data.isoformat(),
                    'time_ago': str(datetime.now() - operation_data)
                }
        
        return jsonify({
            'code': 1,
            'msg': 'success',
            'data': {
                'operations_count': len(frontend_operations),
                'operations': operations_info
            }
        })
        
    except Exception as e:
        return jsonify({'code': 0, 'msg': f'获取前端操作记录失败: {str(e)}'})

@app.route('/api/search/pickup-code/<pickup_code>')
def api_search_by_pickup_code(pickup_code):
    """根据取餐码查询订单"""
    try:
        # 在所有订单中查找匹配的取餐码
        matching_orders = []
        
        for order in orders_db:
            # 从remark字段中提取取餐码
            remark = order.get('remark', '')
            if '取餐码:' in remark:
                order_pickup_code = remark.split('取餐码:')[1].strip()
                if order_pickup_code == pickup_code:
                    matching_orders.append(order)
        
        if matching_orders:
            return jsonify({
                'code': 1,
                'msg': f'找到 {len(matching_orders)} 个匹配的订单',
                'data': matching_orders
            })
        else:
            return jsonify({
                'code': 0,
                'msg': f'未找到取餐码为 {pickup_code} 的订单',
                'data': []
            })
            
    except Exception as e:
        return jsonify({'code': 0, 'msg': f'查询失败: {str(e)}'})

@app.route('/api/search/phone/<phone>')
def api_search_by_phone(phone):
    """根据手机号查询订单"""
    try:
        # 在所有订单中查找匹配的手机号
        matching_orders = []
        
        for order in orders_db:
            if order.get('phone', '') == phone:
                matching_orders.append(order)
        
        if matching_orders:
            return jsonify({
                'code': 1,
                'msg': f'找到 {len(matching_orders)} 个匹配的订单',
                'data': matching_orders
            })
        else:
            return jsonify({
                'code': 0,
                'msg': f'未找到手机号为 {phone} 的订单',
                'data': []
            })
            
    except Exception as e:
        return jsonify({'code': 0, 'msg': f'查询失败: {str(e)}'})

@app.route('/api/system-status')
def api_system_status():
    """获取系统状态"""
    try:
        # 安全处理时间戳
        excel_time_str = None
        if excel_file_modified_time is not None:
            try:
                if hasattr(excel_file_modified_time, 'isoformat'):
                    excel_time_str = excel_file_modified_time.isoformat()
                else:
                    # 如果是float类型的时间戳，转换为datetime
                    from datetime import datetime
                    excel_time_str = datetime.fromtimestamp(excel_file_modified_time).isoformat()
            except Exception:
                excel_time_str = str(excel_file_modified_time)
        
        return jsonify({
            'code': 1,
            'msg': 'success',
            'data': {
                'is_excel_updating': is_excel_updating,
                'excel_file_modified_time': excel_time_str,
                'frontend_operations_count': len(frontend_operations),
                'orders_count': len(orders_db)
            }
        })
    except Exception as e:
        return jsonify({'code': 0, 'msg': f'获取系统状态失败: {str(e)}'})

# 导出函数供main.py使用
def init_app():
    """初始化应用"""
    print("📊 初始化数据读取...")
    
    # 确保Excel文件可写
    ensure_excel_files_writable()
    
    read_excel_orders()
    
    # 启动后台Excel读取线程
    excel_thread = threading.Thread(target=background_excel_reader, daemon=True)
    excel_thread.start()
    print("🔄 后台Excel读取线程已启动，每分钟刷新一次")

def run_app():
    """运行Flask应用"""
    app.run(debug=False, host='0.0.0.0', port=5000) 